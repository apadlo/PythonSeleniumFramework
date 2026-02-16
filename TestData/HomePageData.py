from __future__ import annotations

from pathlib import Path

import openpyxl


class HomePageData:
    test_HomePage_data = [
        {"firstname": "Test", "email": "test@thissheet.com", "gender": "Female"},
        {"firstname": "Zbigniew", "email": "zstonoga@sejm.gov.pl", "gender": "Male"},
    ]

    @staticmethod
    def getTestData(test_case_name):  # noqa: N802
        workbook_path = Path(__file__).resolve().parent / "PythonDemo.xlsx"
        book = openpyxl.load_workbook(workbook_path)
        sheet = book.active
        data = {}

        for i in range(2, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    data[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return [data]
